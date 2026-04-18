#!/usr/bin/env python3
"""VBZ MQTT Live - Meldungsanzeige + Anlagen-Übersicht mit Excel-Import"""

__version__ = "1.0.0"

import sys
import json
import os
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QGroupBox, QLabel, QLineEdit, QPushButton, QTableWidget,
    QTableWidgetItem, QHeaderView, QTextEdit, QSplitter, QCheckBox,
    QSpinBox, QMessageBox, QAbstractItemView, QTabWidget, QComboBox,
    QFileDialog
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QRect, QTimer
from PyQt5.QtGui import QColor, QFont, QIcon, QPixmap, QPainter
import paho.mqtt.client as mqtt


class Settings:
    """Gemeinsame Einstellungen als JSON-Datei neben der EXE.
    Alle Benutzer, die die EXE öffnen können, teilen dieselben Einstellungen."""

    def __init__(self):
        if getattr(sys, "frozen", False):
            base = os.path.dirname(sys.executable)
        else:
            base = os.path.dirname(os.path.abspath(__file__))
        self._path = os.path.join(base, "vbz_mqtt_settings.json")
        self._data = self._load()

    def _load(self) -> dict:
        if os.path.exists(self._path):
            try:
                with open(self._path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                return {}
        return {}

    def _save(self):
        try:
            with open(self._path, "w", encoding="utf-8") as f:
                json.dump(self._data, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"Einstellungen konnten nicht gespeichert werden: {e}")

    def value(self, key: str, default=None):
        return self._data.get(key, default)

    def setValue(self, key: str, value):
        self._data[key] = value
        self._save()


def create_vbz_icon() -> QIcon:
    size = 64
    px = QPixmap(size, size)
    px.fill(QColor("#1888b8"))
    p = QPainter(px)
    p.setPen(Qt.white)
    p.setFont(QFont("Arial", 22, QFont.Bold))
    p.drawText(QRect(0, 0, size, size), Qt.AlignCenter, "VBZ")
    p.end()
    return QIcon(px)


def make_lamp(status: str) -> QLabel:
    label = QLabel("●")
    label.setAlignment(Qt.AlignCenter)
    label.setFont(QFont("Arial", 16))
    color = {"ok": "#22c55e", "error": "#f97316", "offline": "#ef4444"}.get(status, "#ef4444")
    label.setStyleSheet(f"color: {color};")
    tooltip = {"ok": "HEALTH_OK", "error": "Health-Fehler", "offline": "Keine Meldung"}.get(status, "")
    label.setToolTip(tooltip)
    return label


# ---------------------------------------------------------------------------

class MQTTWorker(QThread):
    message_received = pyqtSignal(str, str)
    connection_changed = pyqtSignal(bool, str)

    def __init__(self):
        super().__init__()
        self.client = None

    def connect_broker(self, host, port, username, password, topic):
        self._topic = topic
        self._host = host
        self._port = port
        if self.client:
            try:
                self.client.loop_stop()
                self.client.disconnect()
            except Exception:
                pass
        self.client = mqtt.Client()
        if username:
            self.client.username_pw_set(username, password)
        self.client.on_connect = self._on_connect
        self.client.on_disconnect = self._on_disconnect
        self.client.on_message = self._on_message
        try:
            self.client.connect(host, port, keepalive=60)
            self.client.loop_start()
        except Exception as e:
            self.connection_changed.emit(False, f"Verbindungsfehler: {e}")

    def disconnect_broker(self):
        if self.client:
            try:
                self.client.loop_stop()
                self.client.disconnect()
            except Exception:
                pass

    def _on_connect(self, client, userdata, flags, rc):
        if rc == 0:
            client.subscribe(self._topic)
            self.connection_changed.emit(True, f"Verbunden mit {self._host}:{self._port}, Topic: {self._topic}")
        else:
            codes = {1: "Ungültige Protokollversion", 2: "Ungültige Client-ID",
                     3: "Broker nicht verfügbar", 4: "Falscher Benutzername/Passwort", 5: "Nicht autorisiert"}
            self.connection_changed.emit(False, f"Fehler: {codes.get(rc, f'RC={rc}')}")

    def _on_disconnect(self, client, userdata, rc):
        self.connection_changed.emit(False, "Verbindung getrennt" if rc == 0 else f"Verbindung verloren (RC={rc})")

    def _on_message(self, client, userdata, msg):
        try:
            payload = msg.payload.decode("utf-8", errors="replace")
        except Exception:
            payload = str(msg.payload)
        self.message_received.emit(msg.topic, payload)


# ---------------------------------------------------------------------------

class AnlagenTab(QWidget):
    def __init__(self, device_status: dict, settings: Settings):
        super().__init__()
        self.device_status = device_status
        self.settings = settings
        self.anlagen = []

        # Deaktivierte Anlagen laden
        self.disabled_devices: set = set(self.settings.value("disabled_devices", []))

        self._setup_ui()

        # Gespeicherte Anlagendaten direkt aus Einstellungen laden (kein Dateizugriff nötig)
        saved_anlagen = self.settings.value("anlagen_data", [])
        if saved_anlagen:
            self.anlagen = saved_anlagen
            self._populate_mvu_filter()
            saved_name = self.settings.value("excel_filename", "")
            if saved_name:
                self.lbl_file.setText(f"{saved_name}  ✓ (gespeichert)")
                self.lbl_file.setStyleSheet("color: #1888b8;")
            self.refresh_table()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(6)

        toolbar = QHBoxLayout()
        btn_import = QPushButton("Excel importieren…")
        btn_import.clicked.connect(self.import_excel)
        toolbar.addWidget(btn_import)

        self.lbl_file = QLabel("Keine Datei geladen")
        self.lbl_file.setStyleSheet("color: grey;")
        toolbar.addWidget(self.lbl_file)

        toolbar.addSpacing(20)
        toolbar.addWidget(QLabel("Besitzer (MVU):"))
        self.combo_mvu = QComboBox()
        self.combo_mvu.setMinimumWidth(160)
        self.combo_mvu.addItem("Alle")
        self.combo_mvu.currentTextChanged.connect(self.refresh_table)
        toolbar.addWidget(self.combo_mvu)

        toolbar.addStretch()
        self.lbl_count = QLabel("")
        toolbar.addWidget(self.lbl_count)
        layout.addLayout(toolbar)

        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["Status", "Aktiv", "MVU", "Tech Nr", "Haltestelle"])
        hh = self.table.horizontalHeader()
        hh.setSectionResizeMode(QHeaderView.ResizeToContents)
        hh.setSectionResizeMode(4, QHeaderView.Stretch)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSortingEnabled(True)
        layout.addWidget(self.table)

    def _populate_mvu_filter(self):
        mvus = sorted(set(a["mvu"] for a in self.anlagen if a["mvu"]))
        self.combo_mvu.blockSignals(True)
        self.combo_mvu.clear()
        self.combo_mvu.addItem("Alle")
        for mvu in mvus:
            self.combo_mvu.addItem(mvu)
        self.combo_mvu.blockSignals(False)

    def import_excel(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Excel-Datei öffnen", "", "Excel-Dateien (*.xlsx *.xls)"
        )
        if not path:
            return
        self._load_excel(path)

    def _load_excel(self, path: str):
        try:
            import openpyxl
        except ImportError:
            QMessageBox.critical(self, "Fehler", "Bitte 'openpyxl' installieren:\n  pip install openpyxl")
            return
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active

            col_mvu = col_tech = col_halt = None
            header_row = 0
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                cells = [str(c).strip().lower() if c is not None else "" for c in row]
                if "mvu" in cells:
                    col_mvu  = cells.index("mvu")
                    col_tech = next((j for j, c in enumerate(cells) if "tech" in c), None)
                    col_halt = next((j for j, c in enumerate(cells) if "halt" in c), None)
                    header_row = i + 1
                    break
            if col_mvu is None:
                col_mvu, col_tech, col_halt = 0, 1, 2
                header_row = 1

            self.anlagen = []
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                def cell(idx):
                    return row[idx] if idx is not None and idx < len(row) else None

                mvu_val  = cell(col_mvu)
                tech_val = cell(col_tech)
                halt_val = cell(col_halt)

                if tech_val is None:
                    continue
                tech_str = str(tech_val).strip().split(".")[0]
                if not tech_str.isdigit():
                    continue
                mvu_str = str(mvu_val).strip() if mvu_val else ""
                if "total" in mvu_str.lower():
                    continue

                self.anlagen.append({
                    "mvu":        mvu_str,
                    "tech_nr":    tech_str,
                    "haltestelle": str(halt_val).strip() if halt_val else "",
                })

            self._populate_mvu_filter()

            # Anlagendaten im Settings-File speichern (für alle Benutzer)
            filename = os.path.basename(path)
            self.settings.setValue("anlagen_data",    self.anlagen)
            self.settings.setValue("excel_filename",  filename)

            self.lbl_file.setText(f"{filename}  ✓ (gespeichert)")
            self.lbl_file.setStyleSheet("color: #1888b8;")
            self.refresh_table()

        except Exception as e:
            QMessageBox.critical(self, "Fehler", f"Excel konnte nicht gelesen werden:\n{e}")

    def _make_active_checkbox(self, tech_nr: str) -> QWidget:
        container = QWidget()
        layout = QHBoxLayout(container)
        layout.setContentsMargins(4, 0, 4, 0)
        layout.setAlignment(Qt.AlignCenter)
        chk = QCheckBox()
        chk.setChecked(tech_nr not in self.disabled_devices)
        chk.stateChanged.connect(lambda state, t=tech_nr: self._on_active_changed(t, state))
        layout.addWidget(chk)
        return container

    def _on_active_changed(self, tech_nr: str, state: int):
        if state == Qt.Checked:
            self.disabled_devices.discard(tech_nr)
        else:
            self.disabled_devices.add(tech_nr)
        # Einstellungen speichern (als Liste, für alle Benutzer)
        self.settings.setValue("disabled_devices", list(self.disabled_devices))
        # Zeile sofort durchstreichen/wiederherstellen
        self._apply_strikethrough()

    def _apply_strikethrough(self):
        """Durchstreichen aller deaktivierten Zeilen ohne Tabelle neu aufzubauen."""
        for row in range(self.table.rowCount()):
            tech_item = self.table.item(row, 3)  # Tech Nr Spalte
            if not tech_item:
                continue
            tech_nr = tech_item.text()
            disabled = tech_nr in self.disabled_devices
            font = QFont()
            font.setStrikeOut(disabled)
            color = QColor(180, 180, 180) if disabled else QColor(0, 0, 0)
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item:
                    item.setFont(font)
                    item.setForeground(color)

    def refresh_table(self):
        selected = self.combo_mvu.currentText()
        filtered = [a for a in self.anlagen if selected == "Alle" or a["mvu"] == selected]

        self.table.setSortingEnabled(False)
        self.table.setRowCount(0)
        for anlage in filtered:
            row = self.table.rowCount()
            self.table.insertRow(row)

            status = self.device_status.get(anlage["tech_nr"], "offline")
            self.table.setCellWidget(row, 0, make_lamp(status))
            self.table.setCellWidget(row, 1, self._make_active_checkbox(anlage["tech_nr"]))

            disabled = anlage["tech_nr"] in self.disabled_devices
            font = QFont()
            font.setStrikeOut(disabled)
            color = QColor(180, 180, 180) if disabled else QColor(0, 0, 0)

            for col, key in enumerate(["mvu", "tech_nr", "haltestelle"], start=2):
                item = QTableWidgetItem(anlage[key])
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                item.setFont(font)
                item.setForeground(color)
                if key == "tech_nr":
                    item.setData(Qt.UserRole, int(anlage[key]))
                self.table.setItem(row, col, item)

        self.table.setSortingEnabled(True)
        self.lbl_count.setText(f"{len(filtered)} Anlagen")

    def update_lamps(self):
        selected = self.combo_mvu.currentText()
        filtered = [a for a in self.anlagen if selected == "Alle" or a["mvu"] == selected]
        for row, anlage in enumerate(filtered):
            status = self.device_status.get(anlage["tech_nr"], "offline")
            self.table.setCellWidget(row, 0, make_lamp(status))


# ---------------------------------------------------------------------------

class MeldungenTab(QWidget):
    def __init__(self):
        super().__init__()
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(6)
        layout.addWidget(self._make_filter_panel())
        splitter = QSplitter(Qt.Vertical)
        splitter.addWidget(self._make_table())
        splitter.addWidget(self._make_detail_panel())
        splitter.setSizes([520, 180])
        layout.addWidget(splitter, stretch=1)

    def _make_filter_panel(self):
        box = QGroupBox("Filter")
        outer = QVBoxLayout(box)
        outer.setSpacing(4)

        row1 = QHBoxLayout()
        row1.addWidget(QLabel("SboId:"))
        self.inp_sboid = QLineEdit()
        self.inp_sboid.setPlaceholderText("z.B. 100648")
        self.inp_sboid.setFixedWidth(120)
        self.inp_sboid.setToolTip("Filtert Meldungen, deren Topic diese SboId enthält.\nLeer lassen = alle Topics.")
        row1.addWidget(self.inp_sboid)
        self.chk_health = QCheckBox("Nur Fehler (HEALTH_OK ausblenden)")
        self.chk_health.setChecked(True)
        row1.addWidget(self.chk_health)
        btn_clear = QPushButton("Leeren")
        btn_clear.setFixedWidth(75)
        btn_clear.clicked.connect(self.clear_table)
        row1.addWidget(btn_clear)
        outer.addLayout(row1)

        row2 = QHBoxLayout()
        row2.addWidget(QLabel("Gerätetyp:"))
        self.chk_dcu   = QCheckBox("dcu");   self.chk_dcu.setChecked(True)
        self.chk_du    = QCheckBox("du");    self.chk_du.setChecked(True)
        self.chk_pau   = QCheckBox("pau");   self.chk_pau.setChecked(True)
        self.chk_other = QCheckBox("andere"); self.chk_other.setChecked(False)
        for chk in (self.chk_dcu, self.chk_du, self.chk_pau, self.chk_other):
            row2.addWidget(chk)
        row2.addStretch()
        outer.addLayout(row2)
        return box

    def _make_table(self):
        self.table = QTableWidget()
        self.table.setColumnCount(11)
        self.table.setHorizontalHeaderLabels([
            "Zeitstempel", "Typ", "Topic", "Beschreibung",
            "Health", "Erreichbarkeit", "Aktivierung",
            "Grund", "CPU %", "RAM %", "Disk %",
        ])
        hh = self.table.horizontalHeader()
        hh.setSectionResizeMode(QHeaderView.ResizeToContents)
        hh.setSectionResizeMode(2, QHeaderView.Stretch)
        hh.setSectionResizeMode(3, QHeaderView.Stretch)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.verticalHeader().setVisible(False)
        self.table.setSortingEnabled(True)
        self.table.itemSelectionChanged.connect(self._show_detail)
        return self.table

    def _make_detail_panel(self):
        box = QGroupBox("Details – vollständige Meldung (JSON)")
        lay = QVBoxLayout(box)
        self.detail = QTextEdit()
        self.detail.setReadOnly(True)
        self.detail.setFont(QFont("Courier New", 9))
        lay.addWidget(self.detail)
        return box

    def _find_row(self, topic: str) -> int:
        for r in range(self.table.rowCount()):
            item = self.table.item(r, 0)
            if item and item.data(Qt.UserRole + 1) == topic:
                return r
        return -1

    def add_row(self, topic: str, device_type: str, data: dict, raw: str):
        self.table.setSortingEnabled(False)
        existing = self._find_row(topic)
        row = existing if existing >= 0 else self.table.rowCount()
        if existing < 0:
            self.table.insertRow(row)

        header = data.get("msg_header", {})
        usage  = data.get("usage", {})
        health = data.get("health", "")

        cells = [
            header.get("timestamp", ""), device_type, topic,
            data.get("description", ""), health,
            data.get("reachability", ""), data.get("activation", ""),
            data.get("reason", ""),
            str(usage.get("cpu", "")), str(usage.get("ram", "")), str(usage.get("disk", "")),
        ]
        for col, val in enumerate(cells):
            item = QTableWidgetItem(val)
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            if col == 0:
                item.setData(Qt.UserRole,     raw)
                item.setData(Qt.UserRole + 1, topic)
            self.table.setItem(row, col, item)

        if health in ("HEALTH_OK", ""):
            color = QColor(255, 255, 200)
        elif "WARN" in health or "DEGRADED" in health:
            color = QColor(255, 220, 150)
        else:
            color = QColor(255, 180, 180)
        for col in range(self.table.columnCount()):
            self.table.item(row, col).setBackground(color)

        self.table.setSortingEnabled(True)
        if existing < 0:
            self.table.scrollToBottom()

    def clear_table(self):
        self.table.setRowCount(0)
        self.detail.clear()

    def _show_detail(self):
        row = self.table.currentRow()
        if row < 0:
            return
        item = self.table.item(row, 0)
        if item:
            raw = item.data(Qt.UserRole)
            if raw:
                try:
                    formatted = json.dumps(json.loads(raw), indent=2, ensure_ascii=False)
                except Exception:
                    formatted = raw
                self.detail.setPlainText(formatted)


# ---------------------------------------------------------------------------

class MQTTViewer(QMainWindow):
    def __init__(self):
        super().__init__()
        self.settings = Settings()
        self.worker = MQTTWorker()
        self.worker.message_received.connect(self.handle_message)
        self.worker.connection_changed.connect(self.handle_connection)
        self._total = 0
        self._shown = 0
        self._is_connected = False
        self.device_status = {}
        self._countdown_secs = 0
        self._timer = QTimer()
        self._timer.setInterval(1000)
        self._timer.timeout.connect(self._tick)
        self._setup_ui()
        self._load_settings()

    def _setup_ui(self):
        self.setWindowTitle(f"VBZ MQTT Live  v{__version__}")
        self.setWindowIcon(create_vbz_icon())
        self.setMinimumSize(1300, 750)

        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setSpacing(6)

        main_layout.addWidget(self._make_connection_panel())
        main_layout.addWidget(self._make_countdown_bar())

        self.tabs = QTabWidget()
        self.meldungen_tab = MeldungenTab()
        self.anlagen_tab   = AnlagenTab(self.device_status, self.settings)
        self.tabs.addTab(self.meldungen_tab, "Meldungen")
        self.tabs.addTab(self.anlagen_tab,   "Anlagen")
        main_layout.addWidget(self.tabs, stretch=1)

        self.statusBar().showMessage("Nicht verbunden")

    def _make_countdown_bar(self) -> QWidget:
        bar = QWidget()
        bar.setFixedHeight(36)
        layout = QHBoxLayout(bar)
        layout.setContentsMargins(8, 4, 8, 4)

        self.lbl_countdown = QLabel("Nicht verbunden")
        self.lbl_countdown.setAlignment(Qt.AlignCenter)
        self.lbl_countdown.setFont(QFont("Arial", 11, QFont.Bold))
        self.lbl_countdown.setStyleSheet(
            "color: #555; background: transparent;"
        )
        layout.addWidget(self.lbl_countdown)
        return bar

    def _start_countdown(self):
        self._countdown_secs = 5 * 60
        self._timer.start()
        self._tick()

    def _stop_countdown(self):
        self._timer.stop()
        self.lbl_countdown.setText("Nicht verbunden")
        self.lbl_countdown.setStyleSheet("color: #555; background: transparent;")

    def _tick(self):
        if self._countdown_secs > 0:
            m, s = divmod(self._countdown_secs, 60)
            self.lbl_countdown.setText(
                f"⏱  Noch {m}:{s:02d} min bis alle Anlagen gemeldet haben"
            )
            self.lbl_countdown.setStyleSheet(
                "color: #b45309; background: #fef9c3; border-radius: 6px; padding: 2px 12px;"
            )
            self._countdown_secs -= 1
        else:
            self._timer.stop()
            self.lbl_countdown.setText("✓  Alle Anlagen sollten sich gemeldet haben")
            self.lbl_countdown.setStyleSheet(
                "color: #166534; background: #dcfce7; border-radius: 6px; padding: 2px 12px;"
            )

    def _make_connection_panel(self):
        box = QGroupBox("Broker-Verbindung")
        row = QHBoxLayout(box)

        row.addWidget(QLabel("Host:"))
        self.inp_host = QLineEdit()
        self.inp_host.setFixedWidth(160)
        row.addWidget(self.inp_host)

        row.addWidget(QLabel("Port:"))
        self.inp_port = QSpinBox()
        self.inp_port.setRange(1, 65535)
        self.inp_port.setFixedWidth(75)
        row.addWidget(self.inp_port)

        row.addWidget(QLabel("Benutzer:"))
        self.inp_user = QLineEdit()
        self.inp_user.setFixedWidth(110)
        row.addWidget(self.inp_user)

        row.addWidget(QLabel("Passwort:"))
        self.inp_pass = QLineEdit()
        self.inp_pass.setEchoMode(QLineEdit.Password)
        self.inp_pass.setFixedWidth(110)
        row.addWidget(self.inp_pass)

        row.addWidget(QLabel("Topic:"))
        self.inp_topic = QLineEdit()
        self.inp_topic.setFixedWidth(130)
        row.addWidget(self.inp_topic)

        self.btn_connect = QPushButton("Verbinden")
        self.btn_connect.setFixedWidth(110)
        self.btn_connect.clicked.connect(self._toggle_connection)
        row.addWidget(self.btn_connect)

        return box

    def _load_settings(self):
        self.inp_host.setText(self.settings.value("host", "localhost"))
        self.inp_port.setValue(int(self.settings.value("port", 1883)))
        self.inp_user.setText(self.settings.value("username", ""))
        self.inp_topic.setText(self.settings.value("topic", "#"))

    def _save_settings(self):
        self.settings.setValue("host",     self.inp_host.text().strip())
        self.settings.setValue("port",     self.inp_port.value())
        self.settings.setValue("username", self.inp_user.text().strip())
        self.settings.setValue("topic",    self.inp_topic.text().strip() or "#")

    def _toggle_connection(self):
        if not self._is_connected:
            host = self.inp_host.text().strip()
            if not host:
                QMessageBox.warning(self, "Fehler", "Bitte Host angeben.")
                return
            self._save_settings()
            self.btn_connect.setEnabled(False)
            self.btn_connect.setText("Verbinde…")
            self.worker.connect_broker(
                host, self.inp_port.value(),
                self.inp_user.text().strip(), self.inp_pass.text(),
                self.inp_topic.text().strip() or "#",
            )
        else:
            self.worker.disconnect_broker()

    def handle_connection(self, connected: bool, message: str):
        self._is_connected = connected
        self.btn_connect.setEnabled(True)
        if connected:
            self.btn_connect.setText("Trennen")
            self.statusBar().showMessage(f"✓  {message}  |  Empfangen: 0  |  Angezeigt: 0")
            self._start_countdown()
        else:
            self.btn_connect.setText("Verbinden")
            self.statusBar().showMessage(f"✗  {message}")
            self._stop_countdown()

    @staticmethod
    def _device_type_from_topic(topic: str) -> str:
        known = {"dcu", "du", "pau"}
        for seg in topic.split("/"):
            if seg in known:
                return seg
        return ""

    @staticmethod
    def _tech_nr_from_topic(topic: str):
        known = {"dcu", "du", "pau"}
        parts = topic.split("/")
        for i, seg in enumerate(parts):
            if seg in known and i + 1 < len(parts):
                tech = parts[i + 1].split(":")[0]
                if tech.isdigit():
                    return tech
        return None

    def handle_message(self, topic: str, payload: str):
        self._total += 1
        mt = self.meldungen_tab

        # SboId-Filter
        sboid = mt.inp_sboid.text().strip()
        if sboid and sboid not in topic:
            self._update_status()
            return

        # Gerätetyp-Filter
        device_type = self._device_type_from_topic(topic)
        allowed = set()
        if mt.chk_dcu.isChecked():   allowed.add("dcu")
        if mt.chk_du.isChecked():    allowed.add("du")
        if mt.chk_pau.isChecked():   allowed.add("pau")
        if device_type in {"dcu", "du", "pau"}:
            if device_type not in allowed:
                self._update_status()
                return
        elif not mt.chk_other.isChecked():
            self._update_status()
            return

        # JSON parsen
        try:
            data = json.loads(payload)
        except json.JSONDecodeError:
            self._update_status()
            return

        health   = data.get("health", "")
        tech_nr  = self._tech_nr_from_topic(topic)

        # Deaktivierte Anlage → komplett ignorieren
        if tech_nr and tech_nr in self.anlagen_tab.disabled_devices:
            self._update_status()
            return

        # Anlagen-Status aktualisieren
        if tech_nr:
            new_status = "ok" if health == "HEALTH_OK" else "error"
            if self.device_status.get(tech_nr) != new_status:
                self.device_status[tech_nr] = new_status
                self.anlagen_tab.update_lamps()

        # Health-Filter für Meldungs-Tab
        if mt.chk_health.isChecked() and health == "HEALTH_OK":
            self._update_status()
            return

        is_new = mt._find_row(topic) < 0
        if is_new:
            self._shown += 1
        mt.add_row(topic, device_type, data, payload)
        self._update_status()

    def _update_status(self):
        state = "Verbunden" if self._is_connected else "Getrennt"
        self.statusBar().showMessage(f"{state}  |  Empfangen: {self._total}  |  Angezeigt: {self._shown}")

    def closeEvent(self, event):
        self.worker.disconnect_broker()
        event.accept()


def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    app.setApplicationName("VBZ MQTT Live")
    app.setWindowIcon(create_vbz_icon())
    win = MQTTViewer()
    win.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
